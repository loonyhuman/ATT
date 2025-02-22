import pandas as pd
import os
from openpyxl import Workbook
from typing import List, Tuple
import numpy as np

class DataProcessor:
    @staticmethod
    def zeroing_out(x: str) -> str:
        return '0' + x if len(x) == 1 else x

    def preprocess_data(self, rectangles: List[Tuple[Tuple[int, int], Tuple[int, int]]], kks_list: List[str], units_list: List[str]) -> pd.DataFrame:
        df = pd.DataFrame(rectangles, columns=['first_coor', 'second_coor'])
        df['first_coor_x_template'] = df['first_coor'].apply(lambda x: x[0])
        df['first_coor_y_template'] = df['first_coor'].apply(lambda x: x[1])
        df['second_coor_x_template'] = df['second_coor'].apply(lambda x: x[0])
        df['second_coor_y_template'] = df['second_coor'].apply(lambda x: x[1])
        df['KKS'] = pd.Series(kks_list)
        df['Units'] = pd.Series(units_list)

        mm_pxl = 3.793627
        value_to_add = 0
        df['x_coor_indicator'] = df.apply(lambda row: row['first_coor_x_template'] + value_to_add, axis=1)
        df['y_coor_indicator'] = df.apply(lambda row: row['first_coor_y_template'] + value_to_add, axis=1)

        alphabet = 'ABCDEFGHJKLMNPQRSTUVWXYZ'
        sq_side = 317.475
        df['column_num'] = round(df['x_coor_indicator'] / sq_side)
        df['row_num'] = round(df['y_coor_indicator'] / sq_side)
        df['str_x'] = df.apply(lambda row: str(1 + int(row['column_num'] % sq_side)), axis=1)
        df['str_y'] = df.apply(lambda row: str(alphabet[int(row['row_num'] // len(alphabet))] + alphabet[int(row['row_num'] % len(alphabet))]), axis=1)
        df['str_x'] = df['str_x'].apply(self.zeroing_out)
        df['Coordinate'] = df['str_y'] + '-' + df['str_x']

        return df

    def save_data(self, df: pd.DataFrame, filename: str, output_folder: str):
        filename = os.path.join(output_folder, filename)
        df_to_save = df[['Coordinate', 'KKS', 'Units']]
        df_to_save['SortKey'] = df_to_save['Coordinate'].apply(self.extract_sort_key)
        df_to_save = df_to_save.sort_values(by='SortKey').drop(columns=['SortKey'])

        wb = Workbook()
        ws = wb.active
        for col in df_to_save.columns:
            ws.cell(row=1, column=df_to_save.columns.get_loc(col) + 1, value=col)
        for i, row in df_to_save.iterrows():
            for col, value in row.items():
                ws.cell(row=i + 2, column=df_to_save.columns.get_loc(col) + 1, value=value)
        wb.save(filename)
        print(f"Excel file successfully created and saved as '{filename}'")

    @staticmethod
    def extract_sort_key(coordinate: str) -> Tuple[str, int]:
        parts = coordinate.split('-')
        if len(parts) == 2:
            return parts[0], int(parts[1])
        return coordinate, 0