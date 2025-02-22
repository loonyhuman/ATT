import pandas as pd
import os
from openpyxl import Workbook

# Функция для извлечения букв и чисел из первого столбца
def extract_sort_key(coordinate):
    # Разделяем строку на буквы и числа
    parts = coordinate.split('-')
    if len(parts) == 2:
        return parts[0], int(parts[1])  # Возвращаем буквы и числа как кортеж
    return coordinate, 0  # Если формат неверный, возвращаем исходное значение

def save_data_2B_2009(df: pd.DataFrame, filename: str,  OUTPUT_FOLDER: str):
# Создаем DataFrame
    filename = os.path.join(OUTPUT_FOLDER, filename)
    df_to_save = df.loc[:, ['Coordinate']]

    df_to_save['Name'] = '2B-2009'

    # Создаем новый Excel файл или открываем существующий
    wb = Workbook()
    ws = wb.active

    # Добавляем заголовок таблицы
    for col in df_to_save.columns:  
        ws.cell(row=1, column=df_to_save.columns.get_loc(col) + 1, value=col)

    # Заполняем таблицу данными из DataFrame
    for i, row in df_to_save.iterrows():
        for col, value in row.items():
            ws.cell(row=i+2, column=df_to_save.columns.get_loc(col) + 1, value=value)
            
    wb.save(filename)
    print(f"Excel файл успешно создан и сохранен как '{filename}'")

def save_data_indicator(df: pd.DataFrame, filename: str,  OUTPUT_FOLDER: str):
# Создаем DataFrame
    filename = os.path.join(OUTPUT_FOLDER, filename)
    df_to_save = df.loc[:, ['Coordinate', 'KKS', 'Units']]

    # Добавляем временный столбец для сортировки
    df_to_save['SortKey'] = df_to_save['Coordinate'].apply(extract_sort_key)

    # Сортируем DataFrame по временному столбцу
    df_to_save = df_to_save.sort_values(by='SortKey')

    # Удаляем временный столбец
    df_to_save = df_to_save.drop(columns=['SortKey'])
        
    # Создаем новый Excel файл или открываем существующий
    wb = Workbook()
    ws = wb.active

    # Добавляем заголовок таблицы
    for col in df_to_save.columns:  
        ws.cell(row=1, column=df_to_save.columns.get_loc(col) + 1, value=col)

    # Заполняем таблицу данными из DataFrame
    for i, row in df_to_save.iterrows():
        for col, value in row.items():
            ws.cell(row=i+2, column=df_to_save.columns.get_loc(col) + 1, value=value)
            
    wb.save(filename)
    print(f"Excel файл успешно создан и сохранен как '{filename}'")


def save_data_armatura(df: pd.DataFrame, filename: str, column_names_db: list[str], OUTPUT_FOLDER: str):
# Создаем DataFrame

    # df_to_save = df.loc[:, ['coor', 'KKS', 'Name']+column_names_db]
    filename = os.path.join(OUTPUT_FOLDER, filename)
    df_to_save = df.loc[:, ['Coordinate', 'KKS', 'Name']]

    # Создаем новый Excel файл или открываем существующий
    wb = Workbook()
    ws = wb.active

    # Добавляем заголовок таблицы
    for col in df_to_save.columns:  
        ws.cell(row=1, column=df_to_save.columns.get_loc(col) + 1, value=col)

    # Заполняем таблицу данными из DataFrame
    for i, row in df_to_save.iterrows():
        for col, value in row.items():
            ws.cell(row=i+2, column=df_to_save.columns.get_loc(col) + 1, value=value)

    # Сохраняем Excel-файл
    wb.save('C://Users//user//Desktop//template_programm_divided//Координаты Арматуры.xlsx')
    print("Excel файл успешно создан и сохранен как 'Координаты Арматуры.xlsx'")
    
    
def save_data_switch(df: pd.DataFrame, filename: str, column_names_db: list[str], OUTPUT_FOLDER: str):
# Создаем DataFrame
    filename = os.path.join(OUTPUT_FOLDER, filename)
    df_to_save = df.loc[:, ['Coordinate', 'KKS', 'Name']]

    # Добавляем временный столбец для сортировки
    df_to_save['SortKey'] = df_to_save['Coordinate'].apply(extract_sort_key)

    # Сортируем DataFrame по временному столбцу
    df_to_save = df_to_save.sort_values(by='SortKey')

    # Удаляем временный столбец
    df_to_save = df_to_save.drop(columns=['SortKey'])
        
    # Создаем новый Excel файл или открываем существующий
    wb = Workbook()
    ws = wb.active

    # Добавляем заголовок таблицы
    for col in df_to_save.columns:  
        ws.cell(row=1, column=df_to_save.columns.get_loc(col) + 1, value=col)

    # Заполняем таблицу данными из DataFrame
    for i, row in df_to_save.iterrows():
        for col, value in row.items():
            ws.cell(row=i+2, column=df_to_save.columns.get_loc(col) + 1, value=value)
    wb.save(filename)
    print(f"Excel файл успешно создан и сохранен как '{filename}'")
    
    
def sort_excel():
    pass